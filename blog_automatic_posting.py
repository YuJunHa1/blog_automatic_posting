from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from fake_useragent import UserAgent
from openpyxl import Workbook, load_workbook
from openai import OpenAI
from dotenv import load_dotenv
load_dotenv("info.env")
import time
import random
import os
import pyperclip




def chrome_setting():
    chrome_version = random.randint(118, 122)# 고정된 모바일 버전의 User-Agent 패턴 (숫자만 랜덤하게 변경)
    user_agent = f"Mozilla/5.0 (iPhone; CPU iPhone OS 17_3 like Mac OS X) AppleWebKit/537.36 (KHTML, like Gecko) Version/17.0 Mobile/{chrome_version}.0.0 Safari/537.36"
    options = webdriver.ChromeOptions()
    options.add_argument("accept-language=ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7") #한국어(ko-KR)**를 우선적으로 사용하고, 그다음으로 **영어(en-US, en)**를 선호한다는 것을 나타냅니다.
    options.add_argument("--disable-blink-features=AutomationControlled") #크롬의 셀레니움 자동화 탐지 비활성화
    options.add_argument(f"user-agent={user_agent}") #가짜 User-Agent를 생성하여 쿠팡이 봇을 감지하는 것을 방지
    options.add_experimental_option("excludeSwitches", ["enable-automation"]) #브라우저가 자동화된 환경에서 실행되고 있다는 표시를 숨김
    options.add_experimental_option("useAutomationExtension", False) #브라우저가 자동화된 환경에서 실행되고 있다는 표시를 숨김
    options.add_experimental_option("prefs", {"profile.managed_default_content_setting.images": 2}) #브라우저의 이미지 로딩 차단, 속도 개선
    options.add_experimental_option("detach", True)#창이 닫히지 않고 유지
    driver = webdriver.Chrome(options=options)# 크롬 드라이버 실행
    # 크롤링 방지 탐지 우회
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
    "source": "Object.defineProperty(navigator, 'webdriver', { get: () => undefined })"
    })
    return driver



# 계정 정보 및 검색어 설정
def info_setting(Coupang_id, Coupang_pw, Blog_id, Blog_pw, Blog_write_page, Search_word,File_path, Api_key):   
    global coupang_id
    coupang_id = Coupang_id
    global coupang_pw
    coupang_pw = Coupang_pw
    global coupang_login_page
    coupang_login_page = "https://login.coupang.com/login/login.pang?rtnUrl=https%3A%2F%2Fpartners.coupang.com%2Fapi%2Fv1%2Fpostlogin"
    global blog_id
    blog_id = Blog_id
    global blog_pw
    blog_pw = Blog_pw
    global blog_login_page
    blog_login_page = "https://accounts.kakao.com/login/?continue=https%3A%2F%2Fkauth.kakao.com%2Foauth%2Fauthorize%3Fclient_id%3D3e6ddd834b023f24221217e370daed18%26state%3DaHR0cHM6Ly93d3cudGlzdG9yeS5jb20v%26redirect_uri%3Dhttps%253A%252F%252Fwww.tistory.com%252Fauth%252Fkakao%252Fredirect%26response_type%3Dcode%26auth_tran_id%3DgV7OKYnobRBgoKqmaIF3mJx6CyGt6CNGY9mYaUXMw5EVwsk2YKrO2bLhOjmw%26ka%3Dsdk%252F2.7.3%2520os%252Fjavascript%2520sdk_type%252Fjavascript%2520lang%252Fko-KR%2520device%252FWin32%2520origin%252Fhttps%25253A%25252F%25252Fwww.tistory.com%26is_popup%3Dfalse%26through_account%3Dtrue&talk_login=hidden#login"
    global blog_write_page
    blog_write_page = Blog_write_page
    global search_word
    search_word = Search_word
    global file_path
    file_path = "C:\\blog_automatic_posting\\blog_automatic_posting.xlsx"
    global api_key
    api_key = Api_key
    
def login(driver):
    #아이디 넣기
    driver.get(coupang_login_page)
    login_box = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "login-email-input")))
    login_box.click()
    pyperclip.copy(coupang_id)
    login_box.send_keys(Keys.CONTROL, "v")
    #비밀번호 넣기
    pw_box = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "login-password-input")))
    pw_box.click()
    pyperclip.copy(coupang_pw)
    pw_box.send_keys(Keys.CONTROL, "v")
    #로그인 버튼 누르기, 실패시 계속 실행행
    while(True):
        driver.find_element(By.CSS_SELECTOR, ".login__button--submit").click()
        WebDriverWait(driver, 10).until(EC.url_contains("postlogin"))
        time.sleep(1)
        if not driver.find_elements(By.CSS_SELECTOR, ".login__button--submit"):
            break
    
def search(driver):
    search_box = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ant-input.ant-input-lg")))
    search_box.click()
    pyperclip.copy(search_word)
    search_box.send_keys(Keys.CONTROL, "v")
    driver.find_element(By.CSS_SELECTOR, ".ant-btn.search-button.ant-btn-primary").click()


def get_link(driver, n): #n개의 상품 링크 가져오기
    links = [] # 상품 링크를 담을 리스트
    action = ActionChains(driver)
    for i in range(n):
        #Xpath를 이용해 이번 차례 탐색할 상품 div위로 마우스 이동
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, f"(//div[@class='product-item'])[{i+1}]")))
        action.move_to_element(driver.find_element(By.XPATH, f"(//div[@class='product-item'])[{i+1}]")).perform()
        #마우스가 이동해 링크생성 버튼이 나타나면 해당 버튼을 클릭
        link_button = driver.find_element(By.XPATH, f"(//button[contains(@class, 'btn-generate-link')])[{i+1}]")
        link_button.click()
        #링크 복사 후 리스트에 저장
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".ant-btn.lg.shorten-url-controls-main"))).click()
        time.sleep(1)
        clipboard_text = pyperclip.paste()
        if(clipboard_text[:4] == "http"):
            links.append(clipboard_text)
        driver.back()
    return links



#링크 타고 들어가서 대표 이미지와 베스트 리뷰 가져오기
def get_img_review_title(driver, links):
    img_srcs = []
    best_reviews = []
    titles = []
    for link in links:
        driver.get(link)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, '.rds-button.rds-button--lg.rds-button--fill-blue-normal.rds-button--block.main-product_action__XSlcT'))).click()
        #베스트 리뷰 가져오기기
        img_src = get_img(driver)
        img_srcs.append(img_src)
        best_review = get_review(driver)
        best_reviews.append(best_review)
        title = get_title(driver)
        titles.append(title)
        print(img_srcs)
        print(best_reviews)
        print(titles)
    return img_srcs, best_reviews, titles

def get_img(driver):
    
    try:
        img_src = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".rds-img img"))
        )
        return img_src.get_attribute("src")
    except Exception as e:
            print(f"오류 발생: {e}")
            return None

def get_review(driver):
    best_review = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.CSS_SELECTOR, ".ProductReview_review__article__reviews__text__article__FveJz"))
    )
    return best_review.text

def get_title(driver):
    title = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.CSS_SELECTOR, ".ProductInfo_title__fLscZ"))
    )
    return title.text

    



#링크, 이미지 경로, 리뷰 엑셀에 저장하기
def save_xl(links, img_srcs, best_reviews, titles):
    print("save_xl실행!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
    # 엑셀 파일 경로

    if os.path.exists(file_path):     # 파일이 존재하는지 확인
        wb = load_workbook(file_path) # 파일이 존재하면 열기
    else:
        wb = Workbook() # 파일이 존재하지 않으면 새로 생성

    
    ws = wb.active
    #1행은 속성 값으로 채우기기
    ws.cell(row=1, column=1, value="item_link")
    ws.cell(row=1, column=2, value="img_src")
    ws.cell(row=1, column=3, value="best_review")
    ws.cell(row=1, column=4, value="title")

    #2행부터 각 속성에 맞는 값 채우기
    for i in range(len(links)):
        ws.cell(row=i+2, column=1, value=links[i])
        ws.cell(row=i+2, column=2, value=img_srcs[i])
        ws.cell(row=i+2, column=3, value=best_reviews[i])
        ws.cell(row=i+2, column=4, value=titles[i])

    #중복값 제거
    seen = set()
    rows_to_delete = []

    # 2번째 행부터 데이터 확인 (1번째 행은 헤더)
    for row in range(2, ws.max_row + 1):
        titles = ws.cell(row=row, column=4).value[:8]  # title 컬럼 값 앞 8글자 가져오기, 같은 상품 다른 색상등 거르기 위함

        if titles in seen:  # 이미 존재하면 삭제 리스트에 추가
            rows_to_delete.append(row)
        else:
            seen.add(titles)  # 처음 등장하는 값은 저장

    # 중복 행 삭제 (뒤에서부터 삭제해야 인덱스가 틀어지지 않음)
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)

    #저장
    wb.save(file_path)
    wb.close()

def make_content():
    wb = load_workbook(file_path)
    ws = wb.active
    img_srcs = [cell.value for cell in ws["B"]]
    img_srcs.pop(0)
    best_reviews = [cell.value for cell in ws["C"]]
    best_reviews.pop(0)
    titles = [cell.value for cell in ws["D"]]
    titles.pop(0)
    ws.cell(row=1, column=5, value="contnet")
    # GPT버전 선택
    gpt_version = "gpt-4o-mini"
    # 인증 키 입력
    os.environ["OPENAI_API_KEY"] = api_key
    client = OpenAI()

    contents = [] #생성된 글 담을 리스트

    for i in range(len(best_reviews)):
        query = best_reviews[i] + " 의 내용을 참조해서//" + titles[i] + "제품에 대해 // 20자 미만의 부제목 2개개 뽑아줘."
        sub_title = client.chat.completions.create(
            model=gpt_version,
            messages=[
            {"role": "developer", "content": "You are a helpful assistant."},
            {"role": "user", "content": query}
            ]
        )
        sub_title = sub_title.choices[0].message.content
        sub_titles = sub_title.split("\n") 
        time.sleep(2)


        paragraph = []

        for j in range(len(sub_titles)):
            query = titles[i] + "에 대해서 //" + best_reviews[i] + "//를 참조해서" + sub_titles[j] + "에 대한 내용을 400자 이내로 써줘."
            answer_con = client.chat.completions.create(model=gpt_version,messages=[
            {"role": "developer", "content": "You are a helpful assistant."},
            {"role": "user", "content": query}])
            answer_con = answer_con.choices[0].message.content
            paragraph.append(sub_titles[j] + "\n" + answer_con)
            
        content = "\n".join(paragraph)
        contents.append(content)
        print(str(i+1) + "번째 글 생성 완료")

    #엑셀에 content 저장하기
    i=2 #1열은 속성, 2열부터 값 저장
    for content in contents:
        ws.cell(row=i, column=5, value=content)
        i += 1
    wb.save(file_path)
    wb.close()

def write_blog(driver):

    #엑셀 파일 불러오기
    wb = load_workbook(file_path)
    ws = wb.active

    item_links = [cell.value for cell in ws["A"]] #상품 주소
    item_links.pop(0)
    contents = [cell.value for cell in ws["E"]] #본문 내용
    contents.pop(0)
    img_srcs = [cell.value for cell in ws["B"]] #대표이미지 주소
    img_srcs.pop(0)
    titles = [cell.value for cell in ws["D"]] #상품 이름
    titles.pop(0)

    #로그인
    driver.get(blog_login_page)
    login_box = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "loginId--1")))
    login_box.click()
    pyperclip.copy(blog_id)
    login_box.send_keys(Keys.CONTROL, "v")
    pw_box = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "password--2")))
    pw_box.click()
    pyperclip.copy(blog_pw)
    pw_box.send_keys(Keys.CONTROL, "v")
    driver.find_element(By.CSS_SELECTOR, ".btn_g.highlight.submit").click()

    for i in range(len(contents)):
        time.sleep(2)
        driver.get(blog_write_page)

        #글 이어쓰기 alert창 나오면 새로 쓰기
        try:
            WebDriverWait(driver, 5).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            print("알림창 내용:", alert.text)
            alert.dismiss()
            print("알림창 거절")
        except:
            print("알림창 없음")

        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".textarea_tit"))).click()
        webdriver.ActionChains(driver).send_keys("쿠팡 가성비 " + search_word + " 추천" + titles[i]).perform()

        driver.switch_to.frame("editor-tistory_ifr")
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//p"))).click()
        webdriver.ActionChains(driver).send_keys(f'이 포스팅은 쿠팡 파트너스 활동의 일환으로, 이에 따른 일정액의 수수료를 제공받습니다.\n{item_links[i]}\n').perform()
        time.sleep(5)

        body = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "tinymce")))

        #sub_title1  \n  content1  \n  sub_title2  \n  contnet2 형식으로 되어있는 글을 4개로 쪼개서 각 sub_title에는 <h4> 태그 씌우기
        subtitle_content = contents[i].split("\n")
        # 자바스크립트로 <img> 태그 삽입 (DOM에 직접 넣기)
        img_src = img_srcs[i]
        html = f'<p><img src="{img_src}" alt="쿠팡 이미지"></p><p><h4>{subtitle_content[0]}</h4></p><p>{subtitle_content[1]}</p><p><h4>{subtitle_content[2]}</h4></p><p>{subtitle_content[3]}</p>'

        driver.execute_script("arguments[0].insertAdjacentHTML('beforeend', arguments[1]);", body, html)
        driver.switch_to.default_content()
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".btn.btn-default"))).click()
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "open20"))).click()
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "publish-btn"))).click()




def main(Coupang_id, Coupang_pw, Blog_id, Blog_pw, Blog_write_page, Search_word, Post_num, Api_key):
    driver = chrome_setting()
    info_setting(Coupang_id, Coupang_pw, Blog_id, Blog_pw, Blog_write_page, Search_word, Post_num, Api_key)
    login(driver)
    search(driver)
    links = get_link(driver, int(Post_num)) #가져올 상위 상품 개수
    img_srcs, best_reviews, titles = get_img_review_title(driver, links)
    save_xl(links, img_srcs, best_reviews, titles)
    make_content()
    write_blog(driver)
    print("스크립트 종료")