from openai import OpenAI
from dotenv import load_dotenv
from openpyxl import load_workbook
import time
import os
load_dotenv("info.env")

file_path = "C:\\blog_automatic_posting\\blog_automatic_posting.xlsx"
wb = load_workbook(file_path)
ws = wb.active
img_srcs = [cell.value for cell in ws["B"]]
img_srcs.pop(0)
best_reviews = [cell.value for cell in ws["C"]]
best_reviews.pop(0)
titles = [cell.value for cell in ws["D"]]
titles.pop(0)
ws.cell(row=1, column=5, value="contnet")

# GPT버전 선택
gpt_version = "gpt-4o-mini"
# 인증 키 입력
os.environ["OPENAI_API_KEY"] = os.getenv("OPENAI_API_KEY")
client = OpenAI()

contents = [] #생성된 글 담을 리스트

for i in range(len(best_reviews)):
  
  query = best_reviews[i] + " 의 내용을 참조해서//" + titles[i] + "제품의 구매를 추천하는 글을 쓸 건데 20자 미만의 부제목 2개개 뽑아줘."
  sub_title = client.chat.completions.create(
    model=gpt_version,
    messages=[
      {"role": "developer", "content": "You are a helpful assistant."},
      {"role": "user", "content": query}
    ]
  )
  sub_title = sub_title.choices[0].message.content
  sub_titles = sub_title.split("\n") 
  time.sleep(2)


  paragraph = []

  for j in range(len(sub_titles)):
      query = titles[i] + "에 대해서 //" + best_reviews[i] + "//를 참조해서" + sub_titles[j] + "에 대한 내용을 400자 이내로 써줘."
      answer_con = client.chat.completions.create(model=gpt_version,messages=[
      {"role": "developer", "content": "You are a helpful assistant."},
      {"role": "user", "content": query}])
      answer_con = answer_con.choices[0].message.content
      paragraph.append(sub_titles[j] + "\n" + answer_con)
    
  content = "\n".join(paragraph)
  contents.append(content)
  print(str(i) + "번째 글 생성 완료")

#엑셀에 content 저장하기
i=2 #1열은 속성, 2열부터 값 저장
for content in contents:
   ws.cell(row=i, column=5, value=content)
   i += 1

wb.save(file_path)
wb.close()
print("스크립트 종료")