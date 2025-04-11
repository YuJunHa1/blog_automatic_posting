from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from fake_useragent import UserAgent
from openpyxl import Workbook, load_workbook
from openai import OpenAI
from dotenv import load_dotenv
load_dotenv("info.env")
import time
import random
import os
import pyperclip
#import selenium.webdriver.common.alert import Alert

#alert = Alert(Driver)


blog_id = "qetu5702@gmail.com"
blog_pw = "dbwnsgk7575*"
blog_login_page= "https://accounts.kakao.com/login/?continue=https%3A%2F%2Fkauth.kakao.com%2Foauth%2Fauthorize%3Fclient_id%3D3e6ddd834b023f24221217e370daed18%26state%3DaHR0cHM6Ly93d3cudGlzdG9yeS5jb20v%26redirect_uri%3Dhttps%253A%252F%252Fwww.tistory.com%252Fauth%252Fkakao%252Fredirect%26response_type%3Dcode%26auth_tran_id%3DBhoCei~O9Hgks4VgSLPBq9Hjr5mUGfg1HpxLI99u1UC5iZlTj1cKhQ11Z5LW%26ka%3Dsdk%252F2.7.3%2520os%252Fjavascript%2520sdk_type%252Fjavascript%2520lang%252Fko-KR%2520device%252FWin32%2520origin%252Fhttps%25253A%25252F%25252Fwww.tistory.com%26is_popup%3Dfalse%26through_account%3Dtrue&talk_login=hidden#login"
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)#창이 닫히지 않고 유지
driver = webdriver.Chrome(options=options)# 크롬 드라이버 실행

contents = []
img_srcs = []

#엑셀 파일 불러오기기
file_path = "C:\\blog_automatic_posting\\blog_automatic_posting.xlsx"
wb = load_workbook(file_path)
ws = wb.active

contents = [cell.value for cell in ws["E"]]
contents.pop(0)
img_srcs = [cell.value for cell in ws["B"]]
img_srcs.pop(0)

driver.get(blog_login_page)
login_box = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "loginId--1")))
login_box.click()
pyperclip.copy(blog_id)
login_box.send_keys(Keys.CONTROL, "v")
#비밀번호 넣기
pw_box = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "password--2")))
pw_box.click()
pyperclip.copy(blog_pw)
pw_box.send_keys(Keys.CONTROL, "v")
driver.find_element(By.CSS_SELECTOR, ".btn_g.highlight.submit").click()

time.sleep(5)
#WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".thumb_profile"))).click()
driver.get("https://products-recommendation.tistory.com/manage/newpost?type=post&returnURL=%2Fmanage%2Fposts%2F")

# 잠시 대기 후 alert 감지 및 처리
try:
    WebDriverWait(driver, 2).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    print("알림창 내용:", alert.text)
    alert.dismiss()  # 확인 버튼 클릭
    print("알림창 수락 완료")
except:
    print("알림창 없음")

WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".textarea_tit"))).click()
webdriver.ActionChains(driver).send_keys("제목제목제목목").perform()

driver.switch_to.frame("editor-tistory_ifr")
WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//p"))).click()
webdriver.ActionChains(driver).send_keys("이 포스팅은 쿠팡 파트너스 활동의 일환으로, 이에 따른 일정액의 수수료를 제공받습니다.\nhttps://link.coupang.com/a/cnENdr\n").perform()

body = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "tinymce")))

# 자바스크립트로 <img> 태그 삽입 (DOM에 직접 넣기)
img_url = "http://thumbnail10.coupangcdn.com/thumbnails/remote/292x292ex/image/vendor_inventory/f946/97b84ec7d4b48c3b6fa58a9dbf36fb7c41c152cf18eada1c4f565c6248c1.jpg"
img_html = f'<p><img src="{img_url}" alt="쿠팡 이미지"></p><p>이건 이미지 아래에 들어가는 설명입니다.</p>'

driver.execute_script("arguments[0].insertAdjacentHTML('beforeend', arguments[1]);", body, img_html)
driver.switch_to.default_content()
WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".btn.btn-default"))).click()
WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "open20"))).click()
WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "publish-btn"))).click()

